from collections import defaultdict
import pathlib
import attr
import openpyxl
from clldutils.misc import slug
from pylexibank import Dataset as BaseDataset
from pylexibank import progressbar as pb
from pylexibank import Language
from pylexibank import FormSpec


@attr.s
class CustomLanguage(Language):
    Source = attr.ib(default=None)


class Dataset(BaseDataset):
    dir = pathlib.Path(__file__).parent
    id = "pachechibchan"
    language_class = CustomLanguage
    form_spec = FormSpec(separators="~;,/", missing_data=["∅", "⊘"], first_form_only=True)

    def cmd_makecldf(self, args):
        # add bib
        args.writer.add_sources()
        args.log.info("added sources")

        # add concept
        concepts = {}
        for concept in self.concepts:
            idx = concept["NUMBER"] + "_" + slug(concept["ENGLISH"])
            args.writer.add_concept(
                    ID=idx,
                    Name=concept["ENGLISH"]
                    )
            concepts[concept["ENGLISH"]] = idx
        args.log.info("added concepts")

        # add language
        languages = {}
        sources = defaultdict()
        for language in self.languages:
            args.writer.add_language(
                    ID=language["ID"],
                    Name=language["Name"],
                    Glottocode=language["Glottocode"],
                    Source=language["Source"]
                    )

            languages[language["Name"]] = language["ID"]
            sources[language["Name"]] = language["Source"]
        args.log.info("added languages")

        # read in data
        workbook = openpyxl.load_workbook("raw/Chibchan_Swadesh207_MP.xlsx")
        sheet = workbook.active
        doculects = [cell.value for cell in sheet[1]][1:]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            concept = row[0].value
            for col_idx, cell in enumerate(row[1:], start=1):
                doculect = doculects[col_idx - 1]
                value = cell.value
                print(f"Concept: {concept}, Doculect: {doculect}, Value: {value}")

                if value:
                    args.writer.add_forms_from_value(
                        Language_ID=languages[doculect],
                        Parameter_ID=concepts[concept],
                        Value=value,
                        Source=sources[doculect]
                        )
